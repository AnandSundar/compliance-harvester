"""
Excel Report Generator Module

This module generates an auditor-ready Excel report with three sheets:
1. Summary - Pass/fail counts by service
2. SOC2 - Findings grouped by Trust Service Criterion
3. GDPR - Findings grouped by Article 32 sub-requirement

Uses openpyxl for Excel file generation with professional formatting.
"""

import logging
from datetime import datetime, timezone
from typing import Dict, List, Any
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")


logger = logging.getLogger(__name__)


# Color schemes for status
STATUS_COLORS = {
    "PASS": "C6EFCE",  # Light green
    "FAIL": "FFC7CE",  # Light red
    "MANUAL_REVIEW": "FFEB9C",  # Light yellow
}

SEVERITY_COLORS = {
    "HIGH": "FF6B6B",  # Red
    "MEDIUM": "FFD93D",  # Yellow
    "LOW": "6BCB77",  # Green
}

HEADER_COLOR = "4472C4"  # Blue header


class ExcelReporter:
    """
    Generates Excel compliance reports with multiple sheets.

    This reporter creates a professional auditor-ready report with:
    - Summary sheet with pass/fail counts
    - SOC2 sheet grouped by Trust Service Criteria
    - GDPR sheet grouped by Article 32 sub-requirements
    """

    def __init__(self):
        """Initialize the Excel reporter."""
        self.workbook = Workbook()
        self._setup_styles()

    def _setup_styles(self) -> None:
        """Define styling for the workbook."""
        # Header font
        self.header_font = Font(bold=True, size=12, color="FFFFFF")

        # Header fill
        self.header_fill = PatternFill(
            start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid"
        )

        # Cell alignment
        self.center_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Left alignment
        self.left_align = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

        # Border
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_report(
        self, findings: List[Dict[str, Any]], output_path: str, metadata: Dict[str, Any]
    ) -> None:
        """
        Generate the complete Excel report.

        Args:
            findings: List of all compliance findings
            output_path: Path to save the Excel file
            metadata: Run metadata (account, region, timestamp)
        """
        logger.info(f"Generating Excel report: {output_path}")

        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]

        # Generate all sheets
        self._generate_summary_sheet(findings, metadata)
        self._generate_soc2_sheet(findings)
        self._generate_gdpr_sheet(findings)

        # Save workbook
        self.workbook.save(output_path)
        logger.info(f"Excel report saved: {output_path}")

    def _generate_summary_sheet(
        self, findings: List[Dict[str, Any]], metadata: Dict[str, Any]
    ) -> None:
        """
        Generate Summary sheet with pass/fail counts by service.

        Sheet structure:
        - Header with run metadata
        - Overall pass/fail summary
        - Breakdown by service
        """
        ws = self.workbook.create_sheet("Summary", 0)

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = "Compliance Evidence Summary"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = self.center_align

        # Metadata section
        row = 3
        ws[f"A{row}"] = "Run Information"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        metadata_fields = [
            ("AWS Account", metadata.get("aws_account_id", "N/A")),
            ("Region", metadata.get("region", "N/A")),
            ("Run Timestamp", metadata.get("timestamp", "N/A")),
            ("Tool Version", metadata.get("tool_version", "N/A")),
        ]

        for label, value in metadata_fields:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = str(value)
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Overall summary
        row += 1
        ws[f"A{row}"] = "Overall Summary"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        # Calculate totals
        status_counts = {"PASS": 0, "FAIL": 0, "MANUAL_REVIEW": 0}
        for finding in findings:
            status = finding.get("status", "UNKNOWN")
            if status in status_counts:
                status_counts[status] += 1

        total = sum(status_counts.values())

        row += 1
        headers = ["Status", "Count", "Percentage"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        row += 1
        for status, count in status_counts.items():
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            percentage = (count / total * 100) if total > 0 else 0
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")

            # Apply color
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(
                    start_color=STATUS_COLORS.get(status, "FFFFFF"),
                    end_color=STATUS_COLORS.get(status, "FFFFFF"),
                    fill_type="solid",
                )
                cell.border = self.thin_border

            row += 1

        # Service breakdown
        row += 1
        ws[f"A{row}"] = "Breakdown by Service"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        # Group by service
        service_stats = defaultdict(lambda: {"PASS": 0, "FAIL": 0, "MANUAL_REVIEW": 0})
        for finding in findings:
            check_id = finding.get("check_id", "")
            status = finding.get("status", "UNKNOWN")

            # Determine service from check_id
            if "iam" in check_id or finding.get("resource_id", "").startswith(
                "arn:aws:iam"
            ):
                service = "IAM"
            elif "s3" in check_id or finding.get("resource_id", "").startswith(
                "arn:aws:s3"
            ):
                service = "S3"
            elif "cloudtrail" in check_id or finding.get("resource_id", "").startswith(
                "arn:aws:cloudtrail"
            ):
                service = "CloudTrail"
            elif "config" in check_id or finding.get("resource_id", "").startswith(
                "arn:aws:config"
            ):
                service = "AWS Config"
            else:
                service = "Other"

            if status in service_stats[service]:
                service_stats[service][status] += 1

        row += 1
        headers = ["Service", "PASS", "FAIL", "MANUAL_REVIEW", "Total"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        row += 1
        for service, stats in sorted(service_stats.items()):
            ws.cell(row=row, column=1, value=service)
            ws.cell(row=row, column=2, value=stats["PASS"])
            ws.cell(row=row, column=3, value=stats["FAIL"])
            ws.cell(row=row, column=4, value=stats["MANUAL_REVIEW"])
            total_service = sum(stats.values())
            ws.cell(row=row, column=5, value=total_service)

            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.thin_border

            row += 1

        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _generate_soc2_sheet(self, findings: List[Dict[str, Any]]) -> None:
        """
        Generate SOC2 sheet grouped by Trust Service Criterion.

        Sheet structure:
        - All findings organized by SOC2 criteria
        """
        ws = self.workbook.create_sheet("SOC2")

        # Set column widths
        column_widths = [15, 30, 40, 15, 15, 50]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Group findings by SOC2 criteria
        soc2_findings = defaultdict(list)
        for finding in findings:
            criteria = finding.get("soc2_criteria", [])
            if isinstance(criteria, list):
                for criterion in criteria:
                    soc2_findings[criterion].append(finding)

        # Title
        ws["A1"] = "SOC 2 Trust Service Criteria Findings"
        ws["A1"].font = Font(bold=True, size=14)

        # Headers
        row = 3
        headers = [
            "Criterion",
            "Resource",
            "Check",
            "Status",
            "Severity",
            "Description",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Write findings
        row += 1
        for criterion, criterion_findings in sorted(soc2_findings.items()):
            for finding in criterion_findings:
                ws.cell(row=row, column=1, value=criterion)
                ws.cell(row=row, column=2, value=finding.get("resource_name", "N/A"))
                ws.cell(row=row, column=3, value=finding.get("check_id", ""))
                ws.cell(row=row, column=4, value=finding.get("status", ""))
                ws.cell(row=row, column=5, value=finding.get("severity", ""))
                ws.cell(row=row, column=6, value=finding.get("description", ""))

                # Apply styling
                status = finding.get("status", "")
                severity = finding.get("severity", "")

                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.thin_border
                    cell.alignment = self.left_align

                    # Status color
                    if col == 4:
                        cell.fill = PatternFill(
                            start_color=STATUS_COLORS.get(status, "FFFFFF"),
                            end_color=STATUS_COLORS.get(status, "FFFFFF"),
                            fill_type="solid",
                        )

                    # Severity color
                    if col == 5:
                        cell.fill = PatternFill(
                            start_color=SEVERITY_COLORS.get(severity, "FFFFFF"),
                            end_color=SEVERITY_COLORS.get(severity, "FFFFFF"),
                            fill_type="solid",
                        )

                row += 1

    def _generate_gdpr_sheet(self, findings: List[Dict[str, Any]]) -> None:
        """
        Generate GDPR sheet grouped by Article 32 sub-requirement.

        Sheet structure:
        - All findings organized by GDPR Article 32
        """
        ws = self.workbook.create_sheet("GDPR")

        # Set column widths
        column_widths = [20, 30, 40, 15, 15, 50]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Group findings by GDPR articles
        gdpr_findings = defaultdict(list)
        for finding in findings:
            articles = finding.get("gdpr_articles", [])
            if isinstance(articles, list):
                for article in articles:
                    gdpr_findings[article].append(finding)

        # Title
        ws["A1"] = "GDPR Article 32 Findings"
        ws["A1"].font = Font(bold=True, size=14)

        # Headers
        row = 3
        headers = ["Article", "Resource", "Check", "Status", "Severity", "Description"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Write findings
        row += 1
        for article, article_findings in sorted(gdpr_findings.items()):
            for finding in article_findings:
                ws.cell(row=row, column=1, value=article)
                ws.cell(row=row, column=2, value=finding.get("resource_name", "N/A"))
                ws.cell(row=row, column=3, value=finding.get("check_id", ""))
                ws.cell(row=row, column=4, value=finding.get("status", ""))
                ws.cell(row=row, column=5, value=finding.get("severity", ""))
                ws.cell(row=row, column=6, value=finding.get("description", ""))

                # Apply styling
                status = finding.get("status", "")
                severity = finding.get("severity", "")

                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.thin_border
                    cell.alignment = self.left_align

                    # Status color
                    if col == 4:
                        cell.fill = PatternFill(
                            start_color=STATUS_COLORS.get(status, "FFFFFF"),
                            end_color=STATUS_COLORS.get(status, "FFFFFF"),
                            fill_type="solid",
                        )

                    # Severity color
                    if col == 5:
                        cell.fill = PatternFill(
                            start_color=SEVERITY_COLORS.get(severity, "FFFFFF"),
                            end_color=SEVERITY_COLORS.get(severity, "FFFFFF"),
                            fill_type="solid",
                        )

                row += 1


def generate_excel_report(
    findings: List[Dict[str, Any]], output_path: str, metadata: Dict[str, Any]
) -> None:
    """
    Generate the Excel compliance report.

    Args:
        findings: List of all compliance findings
        output_path: Path to save the Excel file
        metadata: Run metadata (account, region, timestamp)
    """
    reporter = ExcelReporter()
    reporter.generate_report(findings, output_path, metadata)
